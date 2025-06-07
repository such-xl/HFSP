import os
import numpy as np
import pandas as pd
from params import PARAMS

from openpyxl import load_workbook
from scheduling_env.Instance_Generator import Instance_Generator
from scheduling_env.training_env import TrainingEnv
from scheduling_env.model import PPO
from urpic import DMDDQN, DDQN, SHDQN, THDQN
from HMPSAC import HMPSAC

# 方法列表
METHODS = ["RL", "DMDDQN", "DDQN", "THDQN", "HMPSAC"]

num_job = 10
nM = len(METHODS)
s_t = np.zeros((num_job, nM), dtype=float)  # slack_time
U = np.zeros((num_job, nM), dtype=float)  # utilization
makespan = np.zeros((num_job, nM), dtype=float)  # makespan

machine_num = PARAMS["machine_num"]
E_ave = PARAMS["E_ave"]
new_insert = PARAMS["new_insert"]
save_dir = "HFSD/RL/Mig"
path = f"{machine_num}_{E_ave}_{new_insert}.npz"

# 初始化 PPO 并加载模型
ppo = PPO(
    local_state_dim=PARAMS["local_state_dim"],
    local_state_len=PARAMS["local_state_len"],
    global_state_dim=PARAMS["global_state_dim"],
    global_state_len=PARAMS["global_state_len"],
    act_dim=PARAMS["action_dim"],
    a_lr=PARAMS["actor_lr"],
    c_lr=PARAMS["critic_lr"],
    gamma=PARAMS["gamma"],
    lmbda=PARAMS["lmbda"],
    epochs=PARAMS["epochs"],
    eps=PARAMS["eps"],
    device=PARAMS["device"],
    weights=PARAMS["weights"],
    batch_size=PARAMS["batch_size"],
)
dmddqn = DMDDQN(
        local_state_dim=PARAMS["local_state_dim"],
        local_state_len=PARAMS["local_state_len"],
        action_dim=PARAMS["action_dim"],
        weights=PARAMS["weights"],
        gamma=PARAMS["gamma"],
        lr=PARAMS["actor_lr"],
)
ddqn = DDQN(
    PARAMS["local_state_dim"],
    PARAMS["local_state_len"],
    PARAMS["action_dim"],
    weights=PARAMS["weights"],
)
thdqn = THDQN(
    PARAMS["local_state_dim"],
    PARAMS["local_state_len"],
    PARAMS["action_dim"],
    weights=PARAMS["weights"],
)
shdqn = SHDQN(
    PARAMS["local_state_dim"],
    PARAMS["local_state_len"],
    PARAMS["action_dim"],
    weights=PARAMS["weights"],
)
hmpsac = HMPSAC(
        local_state_dim=PARAMS["local_state_dim"],
        local_state_len=PARAMS["local_state_len"],
        global_state_dim=PARAMS["global_state_dim"],
        global_state_len=PARAMS["global_state_len"],
        action_dim=PARAMS["action_dim"],
        weights =  PARAMS["weights"],
        gamma=PARAMS["gamma"],
        a_lr=PARAMS["actor_lr"],
        c_lr=PARAMS["critic_lr"],

)
ppo.load_model(path=f"HFSD/models/ppo_model_RL.pth")
dmddqn.load_model(path=f"HFSD/RL/DMDDQN.pth")
ddqn.load_model(path=f"HFSD/RL/DDQN.pth")
shdqn.load_model(path=f"HFSD/RL/SHDQN.pth")
thdqn.load_model(path=f"HFSD/RL/THDQN.pth")
hmpsac.load_model(path=f"HFSD/RL/HMPSAC.pth")


# 结果记录结构
record = {m: {"slack_time": [], "U": [], "makespan": []} for m in METHODS}


def eval_model(env: TrainingEnv, method: str):
    obs = env.reset(1)
    done = False
    while not done:
        if method == "RL":
            action = ppo.take_action(obs)
            obs, _, done, _ = env.step(action)
        elif method == "DMDDQN":
            action = dmddqn.take_action(obs)
            obs, _, done, _ = env.step(action)
        elif method == "DDQN":
            action = ddqn.take_action(obs)
            obs, _, done, _ = env.step(action)
        # elif method == "SHDQN":
        #     action = shdqn.take_action(obs)
        #     obs, _, done, _ = env.step(action)
        elif method == "THDQN":
            action = thdqn.take_action(obs)
            obs, _, done, _ = env.step(action)
        elif method == "HMPSAC":
            action = hmpsac.take_action(obs)
            obs, _, done, _ = env.step(action)
    record[method]["slack_time"].append(env.compute_slack_time())
    record[method]["U"].append(env.compute_machine_utiliaction())
    record[method]["makespan"].append(env.time_step)
    return env.compute_slack_time(), env.compute_machine_utiliaction(), env.time_step


# 多目标辅助函数
def nondominated_front(points: np.ndarray) -> np.ndarray:
    n = len(points)
    is_pareto = np.ones(n, dtype=bool)
    for i, p in enumerate(points):
        if np.any(np.all(points <= p, axis=1) & np.any(points < p, axis=1)):
            is_pareto[i] = False
    return points[is_pareto]


def generational_distance(approx_pf: np.ndarray, true_pf: np.ndarray) -> float:
    dists = [np.linalg.norm(true_pf - p, axis=1).min() for p in approx_pf]
    return float(np.mean(dists))


def inverted_generational_distance(approx_pf: np.ndarray, true_pf: np.ndarray) -> float:
    dists = [np.linalg.norm(approx_pf - t, axis=1).min() for t in true_pf]
    return float(np.mean(dists))


def ev():
    jobs, arrivals = [], []
    # 1. 先生成所有实例
    for _ in range(num_job):
        jl, al = Instance_Generator(machine_num, E_ave, new_insert)
        jobs.append(jl)
        arrivals.append(al)

    os.makedirs(save_dir, exist_ok=True)

    # 2. 对每个实例评估，并即时保存 npz
    for i in range(num_job):
        # 评估
        envs = [
            TrainingEnv(
                PARAMS["action_dim"],
                machine_num,
                E_ave,
                new_insert,
                jobs[i],
                arrivals[i],
            )
            for _ in METHODS
        ]
        j = 0
        # print(f"\ts_t\tu_m\tms")
        for envs, method in zip(envs, METHODS):
            t, u, m = eval_model(envs, method)
            s_t[i, j] = t
            U[i, j] = u
            makespan[i, j] = m
            j += 1
        # print(f"第 {i + 1} 个实例评估完成：")
            print(f"{method:}\t {t:.2f}\t{u:.4f}\t{m:.2f}")

    # os.makedirs(save_dir, exist_ok=True)
    # np.savez(
    #     os.path.join(save_dir, path),
    #     methods=np.array(METHODS),
    #     slack_time=s_t,
    #     utilization=U,
    #     makespan=makespan,
    # )
    # print(f"所有 {num_job} 次运行数据已保存到 {save_dir}/all_runs.npz")


def spread(approx_pf: np.ndarray, true_pf: np.ndarray) -> float:
    true_sorted = true_pf[np.argsort(true_pf[:, 0])]
    f1_min, f1_max = true_sorted[0], true_sorted[-1]
    ap = approx_pf[np.argsort(approx_pf[:, 0])]
    M = len(ap)
    if M < 2:
        return 0.0
    d = np.linalg.norm(ap[1:] - ap[:-1], axis=1)
    d_bar = d.mean()
    d_f = np.linalg.norm(ap[0] - f1_min)
    d_l = np.linalg.norm(ap[-1] - f1_max)
    return float(
        (d_f + d_l + np.sum(np.abs(d - d_bar))) / (d_f + d_l + (M - 1) * d_bar + 1e-8)
    )


def save_to_excel(save_dir, gd_vals, igd_vals, spr_vals,):
    save_path = save_dir + "/eval.xlsx"
    # 6. 写入 Excel（同前，可按需调整 startrow/startcol）
    book = load_workbook(save_path)

    # 获取每个 sheet 的当前最大行数
    next_row_dict = {}
    for sheet_name in ["GD", "IGD", "Spread"]:
        sheet = book[sheet_name]
        # 获取最大行数（含表头和空行）
        max_row = sheet.max_row
        next_row = max_row + 1  # 下一行
        next_row_dict[sheet_name] = next_row

    # 关闭工作簿（将由 ExcelWriter 管理）
    book.close()
    columns_extended = ["J", "E", "M"] + METHODS
    # 构造新的数据
    gd_df = pd.DataFrame([[new_insert,E_ave,machine_num] + gd_vals], columns=columns_extended)
    igd_df = pd.DataFrame([[new_insert,E_ave,machine_num] + igd_vals], columns=columns_extended)
    spread_df = pd.DataFrame([[new_insert,E_ave,machine_num] + spr_vals], columns=columns_extended)

    # 写入 Excel，不影响原数据
    with pd.ExcelWriter(
        save_path, engine="openpyxl", mode="a", if_sheet_exists="overlay"
    ) as writer:
        gd_df.to_excel(
            writer,
            sheet_name="GD",
            index=False,
            header=False,
            startrow=next_row_dict["GD"] - 1,
            startcol=0,
        )
        igd_df.to_excel(
            writer,
            sheet_name="IGD",
            index=False,
            header=False,
            startrow=next_row_dict["IGD"] - 1,
            startcol=0,
        )
        spread_df.to_excel(
            writer,
            sheet_name="Spread",
            index=False,
            header=False,
            startrow=next_row_dict["Spread"] - 1,
            startcol=0,
        )

    print(f"指标已追加至：{save_path}")

def G_I_S(save_dir,path,METHODS):
    data = np.load(os.path.join(save_dir, path))
    methods_arr = data["methods"]
    slack_mat = data["slack_time"]
    util_mat = data["utilization"]  # 展平所有 runs & methods
    makespan_mat = data["makespan"]

    U_all = util_mat.ravel()
    ms_all = makespan_mat.ravel()
    s_t_all = slack_mat.ravel()

    # 处理三目标
    all_raw = np.stack([1.0 / U_all, ms_all, 1.0 / s_t_all], axis=1)
    mins, maxs = all_raw.min(axis=0), all_raw.max(axis=0)

    def normalize(pts):
        return (pts - mins) / (maxs - mins + 1e-8)

    true_pf = nondominated_front(normalize(all_raw))

    # 5. 计算每个方法指标
    gd_vals, igd_vals, spr_vals = [], [], []
    # print("GD", "IGD", "Spread")
    for j, m in enumerate(METHODS):
        pts = np.stack(
            [1.0 / util_mat[:, j], makespan_mat[:, j], 1.0 / slack_mat[:, j]], axis=1
        )
        # print(np.mean(1.0 / util_mat[:, j]),np.mean(makespan_mat[:, j]),np.mean(1.0 / slack_mat[:, j]),)
        
        norm_pts = normalize(pts)
        apf = nondominated_front(norm_pts)
        gd = generational_distance(apf, true_pf)
        gd = "%.2e" % gd
        igd = inverted_generational_distance(apf, true_pf)
        igd = "%.2e" % igd
        deta = spread(apf, true_pf)
        deta = "%.2e" % deta
        gd_vals.append(gd)
        igd_vals.append(igd)
        spr_vals.append(deta)
        print(f"{m[:2]}:{gd}, {igd}, {deta}")
    return gd_vals,igd_vals,spr_vals
        

if __name__ == "__main__":

    ev()
    gd_vals,igd_vals,spr_vals = G_I_S(save_dir, path, METHODS)
    # save_to_excel(save_dir,gd_vals,igd_vals,spr_vals)
